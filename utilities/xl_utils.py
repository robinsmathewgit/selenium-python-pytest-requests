import openpyxl


def __get_all_header_values(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]

    all_column_headers = []

    for column in range(1, sheet.max_column + 1):
        all_column_headers.append(sheet.cell(row=1, column=int(column)).value)

    return all_column_headers


def __get_column_index(file, sheet_name, column_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    column_index = 0

    for column in range(1, sheet.max_column + 1):
        header_value = str(sheet.cell(row=1, column=int(column)).value)
        if header_value.upper() == column_name.upper():
            column_index = column
            break

    return column_index


def __get_testcase_id_row_index(file, sheet_name, testcase_id):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]

    column = __get_column_index(file, sheet_name, "TestID")

    for row in range(1, sheet.max_row + 1):
        row_value = str(sheet.cell(row=row, column=column).value)
        if row_value.upper() == testcase_id.upper():
            break

    return row


def read_data(file, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    input_value = sheet.cell(row=row_num, column=col_num).value
    return input_value


def write_data(file, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=col_num).value = data
    workbook.save(file)


def get_data(file, sheet_name, testcase_id, column_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]

    column_index = __get_column_index(file, sheet_name, column_name)
    row_index = __get_testcase_id_row_index(file, sheet_name, testcase_id)

    row_value = sheet.cell(row=int(row_index), column=int(column_index)).value
    return row_value
